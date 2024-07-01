import os

from celery_app import make_celery
from celery.schedules import crontab
from run import app
from models import db, QueryHistory
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from api_service import get_vacancies_amount
from datetime import datetime

celery_app = make_celery(app)


@celery_app.task(name='tasks.get_and_save_data')
def get_and_save_data(keywords):
    response = get_vacancies_amount(keywords)

    if 'error' in response.keys():
        error_msg = response['error']
        return error_msg
    try:
        new_query = QueryHistory(vacancies_count=response['success'])
    except KeyError:
        return {'error': 'Invalid response'}
    except Exception as e:
        return {'error': e}
    # TODO: Добавить последний результат vacancies_count в реддис чтобы определять разницу

    db.session.add(new_query)
    db.session.commit()
    return "OK"


@celery_app.task(name='tasks.write_to_excel')
def write_to_excel():
    # Извлечение данных из базы данных
    query = QueryHistory.query.order_by(QueryHistory.datetime).all()
    data = [{'datetime': record.datetime, 'vacancy_count': record.vacancies_count} for record in query]

    # Преобразование данных в DataFrame
    df = pd.DataFrame(data)

    # Добавление столбца 'change'
    df['change'] = df['vacancy_count'].diff().fillna(0).astype(int)

    # Преобразование столбца 'datetime' к нужному формату
    current_date = datetime.utcnow().strftime("%d-%m-%Y")
    formated_datetime = pd.to_datetime(df['datetime'], format='%d-%m-%Y %H:%M')

    df['datetime'] = formated_datetime

    # Запись данных в Excel файл в памяти
    output = f"{current_date}.xlsx"

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Настройка формата заголовка
    wb = load_workbook(output)
    ws = wb.active
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    print(os.getcwd())
    wb.save(output)
    return 'OK'


celery_app.conf.beat_schedule = {
    'get_and_save_data-every-hour': {
        'task': 'tasks.get_and_save_data',
        'schedule': crontab(),  # Каждые часы
        'args': ('junior',),
    },
    # 'write_to_excel-every-hour': {
    #     'task': 'tasks.write_to_excel',
    #     'schedule': crontab(),
    # }
}
